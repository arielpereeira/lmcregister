import openpyxl
import os
from datetime import date
from os.path import exists

##########################################################
###################### VARIÁVEIS #########################
##########################################################

produto = ''
numeracaoTanques = {'Gasolina Aditivada': (1, 4),
                'Gasolina Comum': (1),
                'Diesel': (3)}

workbook = openpyxl.load_workbook('MODELO-LMC.xlsx')

sheet = workbook.active



##########################################################
####################### FUNÇÕES ##########################
##########################################################

# old functions
""" def fNumeroLivro():
    numeroLivro = input("Informe o NÚMERO DO LIVRO: ")
    while numeroLivro == '':
        print("ERRO: Número inválido, tente novamente!\n")
        numeroLivro = input("Informe o NÚMERO DO LIVRO: ")
    cell = sheet.cell(row=1, column=15)
    cell.value = numeroLivro
    print("\n"
          "*******************\n"
          "Número do livro: {}\n"
          "*******************\n".format(numeroLivro))
    return numeroLivro

def fNumeroFolha():
    numeroFolha = input("Informe o NÚMERO DA FOLHA: ")
    while numeroFolha == '':
        print("ERRO: Número inválido, tente novamente!\n")
        numeroFolha = input("Informe o NÚMERO DA FOLHA: ")
    cell = sheet['O2']
    cell.value = numeroFolha
    print("\n"
          "*********************\n"
          "Número da folha: {}\n"
          "*********************\n".format(numeroFolha))
    return numeroFolha """

def fNumeroLivro():
    ano = data[6:]
    numeroLivro = ano
    cell = sheet.cell(row=1, column=15)
    cell.value = numeroLivro
    return numeroLivro
    
def fNumeroFolha():  
    dia = data[0:2]
    mes = data[3:5]  
    numeroFolha = mes+str(dia)
    cell = sheet['O2']
    cell.value = numeroFolha
    return numeroFolha

def fData():
    
    today = date.today()
    hoje = today.strftime("%d/%m/%Y")
    op = input("Você deseja cadastrar a LMC para o dia de hoje ({})? \n"
               "[1] - Sim, irei utilizar a data de hoje.\n"
               "[0] - Não, quero informar uma data. \n"
               "\nInforme um número: ".format(hoje))
    if op == '0':
        data = input("\n\nInforme a data (ex: 01/04/2020): ")
    else:
        data = hoje 
    

    cell = sheet.cell(row=3, column=10) # 2. Data
    cell.value = str(data)
    print("\n"
          "*******************\n"
          "Data: {}\n"
          "*******************\n".format(data))  
    return data

def fEstoqueAbertura(produto):
    if (produto == 'Gasolina Aditivada'):        
        cell = sheet.cell(row=5, column=2) # Tanque 1 do produto
        cell.value = numeracaoTanques[produto][0]
        volEstoqueAbertura = float(input("Informe o volume de {} no tanque {}: ".format(produto,numeracaoTanques[produto][0])))
        cell = sheet.cell(row=6, column=1) # 3.1 Estoque de Abertura
        cell.value = "{} L".format(volEstoqueAbertura)
        print("\n"
          "**************************************\n"
          "Volume de {} no tanque {}: {} L\n"
          "**************************************\n".format(produto,numeracaoTanques[produto][0],volEstoqueAbertura))
        totalEstoqueAbertura = volEstoqueAbertura

        cell = sheet.cell(row=5, column=4) # Tanque 2 do produto
        cell.value = numeracaoTanques[produto][1]
        volEstoqueAbertura = float(input("Informe o volume de {} no tanque {}: ".format(produto,numeracaoTanques[produto][1])))
        cell = sheet.cell(row=6, column=3) # 3.1 Estoque de Abertura
        cell.value = "{} L".format(volEstoqueAbertura)
        print("\n"
          "**************************************\n"
          "Volume de {} no tanque {}: {} L\n"
          "**************************************\n".format(produto,numeracaoTanques[produto][1],volEstoqueAbertura))
        totalEstoqueAbertura += volEstoqueAbertura
        
    else:        
        cell = sheet.cell(row=5, column=2) # Tanque 1 do produto
        cell.value = numeracaoTanques[produto]
        volEstoqueAbertura = float(input("Informe o volume de {} no tanque {}: ".format(produto,numeracaoTanques[produto])))
        cell = sheet.cell(row=6, column=1) # 3.1 Estoque de Abertura
        cell.value = "{} L".format(volEstoqueAbertura)
        print("\n"
          "**************************************\n"
          "Volume de {} no tanque {}: {} L\n"
          "**************************************\n".format(produto,numeracaoTanques[produto],volEstoqueAbertura))
        totalEstoqueAbertura = volEstoqueAbertura

    cell = sheet.cell(row=6, column= 9)
    cell.value = "{} L".format(totalEstoqueAbertura)

    return totalEstoqueAbertura

def fVolumeRecebido():
    row = 9
    totalVolumeRecebido = 0
    op = input("Você comprou {} hoje?\n"
               "[1] - Sim, quero cadastrar uma compra.\n"
               "[0] - Não, quero pular esta etapa. \n"
               "\nInforme um número: ".format(produto))
    while op == '1':
        

        if (op == '1'):
            notaFiscal = input("\nInforme o número da NF: ")
            cell = sheet.cell(row=row, column=3) # Número da NOTA FISCAL
            cell.value = int(notaFiscal)
            print("\n"
                "**************************************\n"
                "Número da NF: {}\n"
                "**************************************\n".format(notaFiscal))

            distribuidor = input("Informe o distribuidor: ")
            cell = sheet.cell(row=row, column=6) # DE: (distribuidor)
            cell.value = str(distribuidor)
            print("\n"
                "**************************************\n"
                "Distribuidor: {}\n"
                "**************************************\n".format(distribuidor))

            volumeRecebido = float(input("Informe o volume recebido de {}: ".format(produto)))
            cell = sheet.cell(row=row, column=13) # 4.2 Volume recebido
            cell.value = "{} L".format(volumeRecebido)
            print("\n"
                "**************************************\n"
                "Volume recebido de {}: {} L\n"
                "**************************************\n".format(produto, volumeRecebido))

                       

            op = input("Quer adicionar outra compra/NF?\n"
                "[1] - Sim, quero cadastrar outra compra.\n"
                "[0] - Não, quero ir para a próxima etapa. \n"
               "\nInforme um número: ") 
            
            row +=1
            totalVolumeRecebido+=volumeRecebido 
        else:
            totalVolumeRecebido = 0
            print("\n"
                "**************************************\n"
                "Volume recebido de {}: {} L\n"
                "**************************************\n".format(produto, volumeRecebido))

    cell = sheet.cell(row=13, column=13) # 4.3 Total recebido 
    cell.value = "{} L".format(totalVolumeRecebido)
    
    return totalVolumeRecebido

def fVolumeVendido():
    row = 18
    totalVolumeVendas = 0
    op = input("Você vendeu {} hoje?\n"
               "[1] - Sim, quero cadastrar uma venda.\n"
               "[0] - Não, quero pular esta etapa. \n"
               "\nInforme um número: ".format(produto))
    while op == '1':

        if(op == '1'):
            if(produto == 'Gasolina Aditivada'):
                numeroTanque = int(input("Informe o número do tanque de {} ({} ou {}): ".format(produto,numeracaoTanques[produto][0],numeracaoTanques[produto][1])))
                cell = sheet.cell(row=row, column=1) # 5.1 Tanque
                cell.value = int(numeroTanque)
                print("\n"
                    "********************\n"
                    "Produto: {}\n"
                    "Tanque: {}\n"
                    "********************\n".format(produto,numeroTanque))
            else:
                numeroTanque = numeracaoTanques[produto]
                cell = sheet.cell(row=row, column=1) # 5.1 Tanque
                cell.value = int(numeroTanque)
                print("\n"
                    "********************\n"
                    "Produto: {}\n"
                    "Tanque: {}\n"
                    "*********************\n".format(produto,numeroTanque))

            numeroBico = int(input("Informe o número do bico do tanque {}: ".format(numeroTanque)))
            cell = sheet.cell(row=row, column=2) # 5.2 Bico
            cell.value = int(numeroBico)
            print("\n"
                "********************\n"
                "Produto: {}\n"
                "Tanque: {}\n"
                "Bico: {}\n"
                "*********************\n".format(produto,numeroTanque,numeroBico))

            volumeFechamento = float(input("Informe o volume de fechamento: "))
            cell = sheet.cell(row=row, column=3) # 5.3 Fechamento
            cell.value = "{} L".format(volumeFechamento)
            print("\n"
                "**************************************\n"
                "Volume de fechamento: {} L\n"
                "**************************************\n".format(volumeFechamento))

            volumeAbertura = float(input("Informe o volume de abertura: "))
            cell = sheet.cell(row=row, column=7) # 5.4 Abertura
            cell.value = "{} L".format(volumeAbertura)
            print("\n"
                "**************************************\n"
                "Volume de fechamento: {} L\n"
                "**************************************\n".format(volumeAbertura))

            volumeAfericao = float(input("Informe o volume de aferição: "))
            cell = sheet.cell(row=row, column=11) # 5.5 Aferições
            cell.value = "{} L".format(volumeAfericao)
            print("\n"
                "**************************************\n"
                "Volume de aferição: {} L\n"
                "**************************************\n".format(volumeAfericao))

            vendasBico = float(volumeFechamento) - float(volumeAbertura)
            cell = sheet.cell(row=row, column=14) # 5.6 Vendas Bico
            cell.value = "{} L".format(vendasBico)

            op = input("Quer adicionar outra venda de {}? \n"
                "[1] - Sim, quero cadastrar outra compra.\n"
            "[0] - Não, quero ir para a próxima etapa. \n"
            "\nInforme um número: ".format(produto)) 
            
            row+=1
            totalVolumeVendas+=vendasBico

        else:
            totalVolumeVendas = 0

    return totalVolumeVendas

def fVendasDoDia():
    valorVendasDoDia = precoCombustivel*vendasDoDia

    return valorVendasDoDia

def fValorAcumuladoMes():
    dia = int(data[0:2]) # precisa ser int para subtrair no path
    mes = data[3:5] 
    ano = data[6:]
    if(dia == 1):
        valorAcumuladoMes = valorVendasDoDia
    else:
        if(dia not in [2, 3, 4, 5, 6, 7, 8, 9]): 
            path = '.\\LMC\\{}\\LMC-{}-{}{}{}.xlsx'.format(produto,prodAbrev,ano,mes,dia-1)
        else:
            path = '.\\LMC\\{}\\LMC-{}-{}{}0{}.xlsx'.format(produto,prodAbrev,ano,mes,dia-1)
        if(fWorkbookExiste(path)):
            wb_read = openpyxl.load_workbook(path)
            sheet_read = wb_read.active
            cell_read = sheet_read.cell(row=35, column=13)
            valor_read = cell_read.value[2:]
            valorAcumuladoMes = float(valor_read) + valorVendasDoDia
        else:
            valorAcumuladoMes = 0


    return valorAcumuladoMes

def fWorkbookExiste(path):
    workbook = exists(path)
    return workbook




##########################################################
####################### EXECUÇÃO #########################
##########################################################


while True: 
    os.system('cls')

    print("*****************************")
    print("\n  Bem-vindo ao LMC Register®")
    print("\n*****************************")

    print("\nO que você deseja fazer?\n")
    print("[1] - Registrar LMC p/ Diesel")
    print("[2] - Registrar LMC p/ Gas. Comum")
    print("[3] - Registrar LMC p/ Gas. Aditivada")
    print("[0] - Sair\n")

    op = input("Informe um número: ")

    while op not in ['1', '2', '3'] or op == None:
        print("Número inválido, informe um número de 1 a 3.\n")
        op = input("Informe um número: ")


    if op == '1':
        produto = "Diesel"
        prodAbrev = "Diesel"
    elif op == '2':
        produto = "Gasolina Comum"
        prodAbrev = 'GasComum'
    elif op == '3':
        produto = "Gasolina Aditivada"
        prodAbrev = "GasAditivada"

    cell = sheet['B3'] # 1. Produto
    cell.value = str(produto) 

    
    while op != '':
        os.system('cls')
        print("***************************************\n")
        print(" [Registrando LMC p/ {}]\n".format(produto))
        print("***************************************")
        print("\nAVISO: Preste muita atenção na hora de registrar as informações!\n")
        op = input("Pressione ENTER para começar...")  

    print("____________________________________________________________________\n")

    print("1) DATA\n")
    data = fData()       
    numeroLivro = fNumeroLivro()
    numeroFolha = fNumeroFolha()



    print("____________________________________________________________________\n")    

    print("2) ESTOQUE DE ABERTURA\n")
    estoqueAbertura = fEstoqueAbertura(produto)

    print("____________________________________________________________________\n")

    print("3) VOLUME RECEBIDO\n")
    volumeRecebido = fVolumeRecebido()

    volumeDisponivel = float(estoqueAbertura) + volumeRecebido
    cell = sheet.cell(row=14, column=13) # 4.4 Volume Disponível (3.1 + 4.3)
    cell.value = "{} L".format(volumeDisponivel)
        

    print("____________________________________________________________________\n")

    print("4) VOLUME VENDIDO\n")
    vendasDoDia = fVolumeVendido()

    cell = sheet.cell(row=26, column=14) # 5.7 Vendas do Dia
    cell.value = "{} L".format(vendasDoDia)

    estoqueEscritural = float(volumeDisponivel) - float(vendasDoDia)
    cell = sheet.cell(row=27, column=14) # 6. Estoque Escritural (4.4 Volume Disponível - 5.7 Vendas do Dia)
    cell.value = "{} L".format(estoqueEscritural)

    print("____________________________________________________________________\n")

    print("5) PREÇO DO COMBUSTÍVEL\n")
    precoCombustivel = float(input("Informe o preço de {}: ".format(produto)))
    cell = sheet.cell(row=26,column=7) # Preço do combustível
    cell.value = float(precoCombustivel)  
    print("\n"
        "**************************************\n"
        "Preço de {}: {} L\n"
        "**************************************\n".format(produto,precoCombustivel)) 


    estoqueFechamento = float(vendasDoDia)
    cell = sheet.cell(row=29, column=5) # 7. Estoque de Fechamento
    cell.value = "{} L".format(estoqueFechamento)
        

    perdasGanhos = float(estoqueEscritural) - float(estoqueFechamento)
    cell = sheet.cell(row=29, column=13) # 8. Perdas e Ganhos
    cell.value = "{} L".format(perdasGanhos)

    cell = sheet.cell(row=32, column=13)
    cell.value = "{} L".format(estoqueFechamento) # 9.1 Total do estoque

    valorVendasDoDia = fVendasDoDia()
    cell = sheet.cell(row=35, column=5) # 10.1 Valor vendas do dia
    cell.value = "R$ {}".format(valorVendasDoDia)

    valorAcumuladoMes = fValorAcumuladoMes()
    cell = sheet.cell(row=35, column=13) # 10.2 Valor acumulado do mês
    cell.value = "R$ {}".format(valorAcumuladoMes)


    file = "LMC-{}-{}{}.xlsx".format(prodAbrev,numeroLivro,numeroFolha)
    path = '.\\LMC\\{}\\{}'.format(produto,file)

    
    print("LMC registrada com sucesso.\n"
          "Salvo em: {}".format(path))
    workbook.save(path)
    
    print("\n###########################################")

    print("\nO que você deseja fazer?\n")
    print("[1] - Registrar nova LMC")
    print("[0] - Sair\n")
    if int(input("Informe um número: ")) != 1:
        break

